"""
Erzeugt eine Beispiel-Excel mit derselben Struktur wie die echte Produktiv-Datei,
aber ausschließlich frei erfundenen Dummy-Daten. Keine echten Kunden/Adressen.

    python3 sample_data/generate_sample_excel.py

Schreibt sample_data/beispiel_bauauftraege.xlsx.
"""
import os
import random
from datetime import datetime, timedelta

import openpyxl
from openpyxl.utils import column_index_from_string

random.seed(42)

# Spalten-Layout wie in backend/excel_reader.py (_COLS_BASE)
COLS = {
    "A": "kls_id", "B": "eingangsdatum", "E": "genehmigung",
    "F": "genehm_von", "G": "genehm_bis", "J": "ansprechpartner",
    "R": "strasse", "S": "hausnummer", "T": "hausnummer_z", "U": "name",
    "W": "telefon", "X": "kommentar", "Y": "ordner_link",
    "Z": "tb_termin", "AA": "tb_team", "AC": "tb_kommentar",
    "AD": "inst_termin", "AE": "inst_team", "AK": "inst_kommentar",
    "AO": "status", "AR": "kw_bau", "AT": "kw_tel",
}
# Wartegrund-Spalte je Stadt (wie SHEETS in excel_reader.py)
WG_COL = {"Stadt A": "AN", "Stadt B": "AN", "Stadt C": "AM",
          "Stadt D": "AL", "Stadt E": "AN", "Stadt F": "AN", "Stadt G": "AN"}

STAEDTE = ["Stadt A", "Stadt B", "Stadt C", "Stadt D", "Stadt E", "Stadt F", "Stadt G"]

STATUS = [
    "offen", "Tiefbau terminiert", "TB abgeschlossen",
    "Installation abgeschlossen", "Auftrag abgeschlossen",
    "Prozess abgeschlossen", "Storno",
]
STRASSEN = ["Musterstraße", "Beispielweg", "Lindenallee", "Bahnhofstraße",
            "Gartenweg", "Am Hang", "Ringstraße", "Schulstraße", "Feldweg"]
VORNAMEN = ["Anna", "Ben", "Clara", "David", "Ella", "Felix", "Greta", "Hannes",
            "Ida", "Jonas", "Klara", "Lukas", "Mia", "Noah", "Olivia", "Paul"]
NACHNAMEN = ["Müller", "Schmidt", "Weber", "Becker", "Fischer", "Wagner",
             "Koch", "Richter", "Klein", "Wolf", "Neumann", "Schwarz"]
TEAMS = ["Team 1", "Team 2", "Team 3"]
WARTEGRUENDE = ["", "", "", "Genehmigung ausstehend", "Kunde nicht erreichbar",
                "Material fehlt"]


def _cell(ws, row, col_letter, value):
    ws.cell(row=row, column=column_index_from_string(col_letter), value=value)


def _fuelle_stadt(ws, stadt, n):
    ws.cell(row=1, column=1, value=f"Bauaufträge {stadt} (Dummy-Daten)")
    ws.cell(row=2, column=1, value="Kopfzeile — echte Datei hat hier weitere Spaltenüberschriften")
    heute = datetime(2026, 6, 1)
    for i in range(n):
        r = i + 3  # Daten ab Zeile 3
        status = random.choice(STATUS)
        eingang = heute - timedelta(days=random.randint(10, 200))
        tb = eingang + timedelta(days=random.randint(5, 40))
        inst = tb + timedelta(days=random.randint(3, 30))
        kw = tb.isocalendar().week
        _cell(ws, r, "A", 10000 + i + STAEDTE.index(stadt) * 100)
        _cell(ws, r, "B", eingang.strftime("%d.%m.%Y"))
        _cell(ws, r, "E", random.choice(["Ja", "Nein"]))
        _cell(ws, r, "F", eingang.strftime("%d.%m.%Y"))
        _cell(ws, r, "G", (eingang + timedelta(days=90)).strftime("%d.%m.%Y"))
        # Stadt A wird per Spalte J in Intern/Extern gesplittet
        _cell(ws, r, "J", random.choice(["Intern", "Extern"]) if stadt == "Stadt A" else "Intern")
        _cell(ws, r, "R", random.choice(STRASSEN))
        _cell(ws, r, "S", str(random.randint(1, 120)))
        _cell(ws, r, "T", random.choice(["", "", "A", "B"]))
        _cell(ws, r, "U", f"{random.choice(VORNAMEN)} {random.choice(NACHNAMEN)}")
        _cell(ws, r, "W", f"0151{random.randint(1000000, 9999999)}")
        _cell(ws, r, "X", random.choice(["", "", "Rückruf gewünscht", "Zugang über Hinterhof"]))
        _cell(ws, r, "Y", "")
        _cell(ws, r, "Z", tb.strftime("%d.%m.%Y %H:%M"))
        _cell(ws, r, "AA", random.choice(TEAMS))
        _cell(ws, r, "AD", inst.strftime("%d.%m.%Y %H:%M"))
        _cell(ws, r, "AE", random.choice(TEAMS))
        _cell(ws, r, WG_COL[stadt], random.choice(WARTEGRUENDE) if status == "offen" else "")
        _cell(ws, r, "AO", status)
        if status in ("TB abgeschlossen", "Auftrag abgeschlossen", "Prozess abgeschlossen"):
            _cell(ws, r, "AR", kw)  # KW Tiefbau fertig
        if status in ("Installation abgeschlossen", "Prozess abgeschlossen"):
            _cell(ws, r, "AT", kw + random.randint(1, 3))  # KW Telekom fertig


def _fuelle_ansprechpartner(ws):
    ws.append(["Gruppe", "Name", "E-Mail", "Funktion", "Telefon"])
    daten = [
        ("Tiefbau", "Anna Müller", "anna.mueller@example.com", "Bauleitung", "0151111111"),
        ("Tiefbau", "Ben Schmidt", "ben.schmidt@example.com", "Polier", "0151222222"),
        ("Installation", "Clara Weber", "clara.weber@example.com", "Technik", "0151333333"),
        ("Vertrieb", "David Becker", "david.becker@example.com", "Kundenkontakt", "0151444444"),
    ]
    for gruppe, name, mail, funktion, tel in daten:
        ws.append([gruppe, name, mail, funktion, tel])


def _fuelle_verfuegbarkeit(ws):
    ws.append(["Name", "Team", "Von", "Bis", "Grund"])
    daten = [
        ("Ben Schmidt", "Team 1", "10.06.2026", "20.06.2026", "Urlaub"),
        ("Clara Weber", "Team 2", "15.06.2026", "16.06.2026", "Fortbildung"),
        ("Felix Wolf", "Team 3", "01.07.2026", "14.07.2026", "Urlaub"),
    ]
    for row in daten:
        ws.append(list(row))


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for stadt in STAEDTE:
        ws = wb.create_sheet(stadt)
        _fuelle_stadt(ws, stadt, random.randint(8, 14))
    _fuelle_ansprechpartner(wb.create_sheet("Ansprechpartner"))
    _fuelle_verfuegbarkeit(wb.create_sheet("Verfügbarkeit"))

    out = os.path.join(os.path.dirname(__file__), "beispiel_bauauftraege.xlsx")
    wb.save(out)
    print("geschrieben:", out)


if __name__ == "__main__":
    main()
