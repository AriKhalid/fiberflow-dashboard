import os
import re
import openpyxl
from datetime import datetime, date

EXCEL_PATH = os.environ.get(
    "EXCEL_PATH",
    os.path.join(os.path.dirname(__file__), "..", "sample_data", "beispiel_bauauftraege.xlsx"),
)

_COLS_BASE = {
    'kls_id':           'A',
    'eingangsdatum':    'B',
    'genehmigung':      'E',
    'genehm_von':       'F',
    'genehm_bis':       'G',
    'ansprechpartner':  'J',
    'strasse': 'R', 'hausnummer': 'S', 'hausnummer_z': 'T', 'name': 'U', 'telefon': 'W',
    'kommentar': 'X',
    'ordner_link': 'Y',
    'tb_termin': 'Z', 'tb_team': 'AA',
    'tb_kommentar': 'AC',
    'inst_termin': 'AD', 'inst_team': 'AE',
    'inst_kommentar': 'AK',
    'status': 'AO',
}

# Wartegrund-Text steht je nach Sheet in einer anderen Spalte
SHEETS = {
    'Stadt A': {**_COLS_BASE, 'wartegrund': 'AN'},  # AN=39
    'Stadt B': {**_COLS_BASE, 'wartegrund': 'AN'},  # AN=39
    'Stadt C':     {**_COLS_BASE, 'wartegrund': 'AM'},  # AM=38
    'Stadt D':       {**_COLS_BASE, 'wartegrund': 'AL'},  # AL=37
    'Stadt E':      {**_COLS_BASE, 'wartegrund': 'AN'},  # AN=39 (Struktur wie Stadt A)
    'Stadt F': {**_COLS_BASE, 'wartegrund': 'AN'},  # AN=39 (Struktur wie Stadt A)
}

# Wartegrund-Spaltenindex je Stadt (für load_kpi_status)
_WG_IDX = {
    'Stadt A': 39,
    'Stadt B': 39,
    'Stadt C':     38,
    'Stadt D':       37,
    'Stadt E':      39,
    'Stadt F': 39,
}


def col_to_idx(col: str | None) -> int | None:
    if col is None:
        return None
    col = col.upper()
    if len(col) == 1:
        return ord(col) - ord('A')
    return (ord(col[0]) - ord('A') + 1) * 26 + (ord(col[1]) - ord('A'))


def cell(row, col_letter: str | None):
    if col_letter is None:
        return None
    idx = col_to_idx(col_letter)
    if idx is None or idx >= len(row):
        return None
    return row[idx].value


def parse_date_range(val) -> tuple:
    """Parst ein Datum oder einen Bereich 'DD.MM.YYYY - DD.MM.YYYY'.
    Gibt (start_datetime, end_datetime_or_None) zurück."""
    if val is None:
        return None, None
    s = str(val).strip()
    if ' - ' in s:
        parts = s.split(' - ', 1)
        start = parse_datetime(parts[0].strip())
        end   = parse_datetime(parts[1].strip())
        return start, end
    return parse_datetime(val), None


def parse_datetime(val) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    val = str(val).strip().replace('\n', ' ')
    for fmt in ('%d.%m.%Y %H:%M:%S', '%d.%m.%Y %H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d.%m.%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(val, fmt)
        except Exception:
            continue
    return None


def full_address(t: dict) -> str:
    hnr = str(t.get('hausnummer') or '').strip()
    hnr_z = str(t.get('hausnummer_z') or '').strip()
    hausnummer_full = f"{hnr}{hnr_z}".strip()
    if hausnummer_full:
        return f"{t['strasse']} {hausnummer_full}, {t['stadt']}"
    return f"{t['strasse']}, {t['stadt']}"


_STATUS_REIHENFOLGE = [
    "offen",
    "Wartegrund",
    "Tiefbau terminiert",
    "TB abgeschlossen",
    "TB bereits abgeschlossen",
    "Installation Problem",
    "Installation abgeschlossen",
    "Auftrag abgeschlossen",
    "Prozess abgeschlossen",
    "Storno",
]

def load_kpi_status() -> dict:
    """Zählt rohe Status-Werte pro Stadt. Stadt A wird per Spalte J in Stadt A Intern / Stadt A Extern aufgeteilt."""
    if not os.path.exists(EXCEL_PATH):
        return {}
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, keep_vba=False, data_only=True)

    staedte_label = ["Stadt A Intern", "Stadt A Extern", "Stadt B", "Stadt D", "Stadt C", "Stadt E", "Stadt F"]
    result = {s: {k: 0 for k in _STATUS_REIHENFOLGE} for s in staedte_label}

    def _kpi_status(row, wg_idx):
        wg   = str(row[wg_idx] or "").strip() if len(row) > wg_idx else ""
        stat = str(row[40] or "").strip() if len(row) > 40 else ""
        if stat.lower() == "offen" and wg:
            return "Wartegrund"
        return stat

    # Stadt A: per Spalte J (Index 9) aufteilen
    if "Stadt A" in wb.sheetnames:
        for row in wb["Stadt A"].iter_rows(min_row=3, values_only=True):
            ansp = str(row[9] or "").strip() if len(row) > 9 else ""
            status = _kpi_status(row, _WG_IDX["Stadt A"])
            if not status:
                continue
            label = "Stadt A Extern" if "Extern" in ansp else "Stadt A Intern"
            if status in result[label]:
                result[label][status] += 1

    # Restliche Städte
    for sheet, label in [("Stadt B", "Stadt B"), ("Stadt D", "Stadt D"), ("Stadt C", "Stadt C"), ("Stadt E", "Stadt E"), ("Stadt F", "Stadt F")]:
        if sheet not in wb.sheetnames:
            continue
        for row in wb[sheet].iter_rows(min_row=3, values_only=True):
            status = _kpi_status(row, _WG_IDX[sheet])
            if status and status in result[label]:
                result[label][status] += 1

    wb.close()

    gesamt = {k: 0 for k in _STATUS_REIHENFOLGE}
    for sd in result.values():
        for k, v in sd.items():
            gesamt[k] += v

    return {
        "staedte": staedte_label,
        "status_reihenfolge": _STATUS_REIHENFOLGE,
        "daten": result,
        "gesamt": gesamt,
    }


def load_kpi_wochentlich() -> dict:
    if not os.path.exists(EXCEL_PATH):
        return {"bau": {}, "telekom": {}}
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, keep_vba=False, data_only=True)

    staedte = ["Stadt A", "Stadt B", "Stadt D", "Stadt C", "Stadt E", "Stadt F"]
    bau: dict     = {}  # {kw: {stadt: count}}
    telekom: dict = {}  # {kw: {stadt: count}}

    for stadt in staedte:
        if stadt not in wb.sheetnames:
            continue
        ws = wb[stadt]
        rows = list(ws.rows)
        for row in rows[2:]:
            kw_bau = row[43].value if len(row) > 43 else None     # AR
            kw_tel = row[45].value if len(row) > 45 else None     # AT
            if kw_bau:
                try:
                    kw = int(kw_bau)
                    bau.setdefault(kw, {}).setdefault(stadt, 0)
                    bau[kw][stadt] += 1
                except (ValueError, TypeError):
                    pass
            if kw_tel:
                try:
                    kw = int(kw_tel)
                    telekom.setdefault(kw, {}).setdefault(stadt, 0)
                    telekom[kw][stadt] += 1
                except (ValueError, TypeError):
                    pass

    wb.close()
    return {"bau": bau, "telekom": telekom}


def load_ansprechpartner() -> list[dict]:
    if not os.path.exists(EXCEL_PATH):
        return []
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, keep_vba=False, data_only=True)
    if 'Ansprechpartner' not in wb.sheetnames:
        wb.close()
        return []
    ws = wb['Ansprechpartner']
    rows = list(ws.rows)
    wb.close()

    result = []
    current_gruppe = None
    for row in rows[1:]:
        gruppe_val   = row[0].value if len(row) > 0 else None
        name_val     = row[1].value if len(row) > 1 else None
        email_val    = row[2].value if len(row) > 2 else None
        funktion_val = row[3].value if len(row) > 3 else None
        tel_val      = row[4].value if len(row) > 4 else None

        if gruppe_val:
            current_gruppe = str(gruppe_val).strip()
        if not name_val:
            continue
        result.append({
            'gruppe':   current_gruppe or '',
            'name':     str(name_val).strip(),
            'email':    str(email_val).strip() if email_val else '',
            'funktion': str(funktion_val).strip() if funktion_val else '',
            'telefon':  str(tel_val).strip() if tel_val else '',
        })
    return result


def load_kpi_nach_kw() -> dict:
    """Zählt 'Prozess abgeschlossen' pro KW und Stadt. Stadt A wird per Spalte J aufgeteilt."""
    if not os.path.exists(EXCEL_PATH):
        return {}

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, keep_vba=False, data_only=True)

    # kw -> {label: count}
    LABELS = ["Stadt A Intern", "Stadt A Extern", "Stadt B", "Stadt D", "Stadt C", "Stadt E", "Stadt F", "Stadt G"]
    daten: dict = {}  # {kw: {label: count}}

    def _add(kw_val, label):
        try:
            kw = int(kw_val)
        except (ValueError, TypeError):
            return
        if kw not in daten:
            daten[kw] = {l: 0 for l in LABELS}
        if label in daten[kw]:
            daten[kw][label] += 1

    # Stadt A: split per Spalte J
    if "Stadt A" in wb.sheetnames:
        for row in wb["Stadt A"].iter_rows(min_row=3, values_only=True):
            if len(row) <= 43:
                continue
            status = str(row[40] or "").strip()
            if status != "Prozess abgeschlossen":
                continue
            kw   = row[43]
            ansp = str(row[9] or "").strip()
            label = "Stadt A Extern" if "Extern" in ansp else "Stadt A Intern"
            _add(kw, label)

    # Restliche Städte
    for sheet, label in [("Stadt B", "Stadt B"), ("Stadt D", "Stadt D"), ("Stadt C", "Stadt C"), ("Stadt E", "Stadt E"), ("Stadt F", "Stadt F"), ("Stadt G", "Stadt G")]:
        if sheet not in wb.sheetnames:
            continue
        for row in wb[sheet].iter_rows(min_row=3, values_only=True):
            if len(row) <= 43:
                continue
            status = str(row[40] or "").strip()
            if status != "Prozess abgeschlossen":
                continue
            _add(row[43], label)

    wb.close()

    # Gesamt pro KW berechnen
    ergebnis = {}
    for kw in sorted(daten.keys()):
        zeile = dict(daten[kw])
        zeile["gesamt"] = sum(zeile.values())
        ergebnis[kw] = zeile

    return {"labels": LABELS, "daten": ergebnis}


def load_verfuegbarkeit() -> list[dict]:
    if not os.path.exists(EXCEL_PATH):
        return []
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, keep_vba=False, data_only=True)
    if 'Verfügbarkeit' not in wb.sheetnames:
        wb.close()
        return []
    ws = wb['Verfügbarkeit']
    rows = list(ws.rows)
    wb.close()

    result = []
    for row in rows[1:]:
        name_val  = row[0].value if len(row) > 0 else None
        team_val  = row[1].value if len(row) > 1 else None
        von_val   = row[2].value if len(row) > 2 else None
        bis_val   = row[3].value if len(row) > 3 else None
        grund_val = row[4].value if len(row) > 4 else None

        if not team_val:
            continue
        von_dt = parse_datetime(von_val)
        bis_dt = parse_datetime(bis_val)
        if not von_dt or not bis_dt:
            continue

        result.append({
            'name':  str(name_val).strip() if name_val else '',
            'team':  str(team_val).strip(),
            'von':   von_dt.strftime('%Y-%m-%d'),
            'bis':   bis_dt.strftime('%Y-%m-%d'),
            'grund': str(grund_val).strip() if grund_val else '',
        })
    return result


def load_alle_auftraege() -> list[dict]:
    if not os.path.exists(EXCEL_PATH):
        return []

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, keep_vba=False, data_only=True)
    result = []

    for stadt, cols in SHEETS.items():
        if stadt not in wb.sheetnames:
            continue
        ws = wb[stadt]
        rows = list(ws.rows)
        if len(rows) < 3:
            continue

        for row in rows[2:]:
            strasse = cell(row, cols['strasse'])
            if not strasse:
                continue

            tb_dt, tb_bis_dt = parse_date_range(cell(row, cols['tb_termin']))
            inst_dt      = parse_datetime(cell(row, cols['inst_termin']))
            eingang_dt   = parse_datetime(cell(row, cols.get('eingangsdatum')))
            genehm_von   = parse_datetime(cell(row, cols.get('genehm_von')))
            genehm_bis   = parse_datetime(cell(row, cols.get('genehm_bis')))
            genehm_raw   = str(cell(row, cols.get('genehmigung')) or '').strip().lower()
            genehmigung  = genehm_raw in ('ja', 'yes', '1', 'true')

            wg_text    = str(cell(row, cols.get('wartegrund')) or '').strip()
            status_raw = str(cell(row, cols['status']) or '').strip()
            status_val = 'Wartegrund' if status_raw.lower() == 'offen' and wg_text else status_raw

            kls_raw = cell(row, cols.get('kls_id'))
            result.append({
                'kls_id':        str(int(kls_raw)) if kls_raw and str(kls_raw).strip() != '' else '',
                'stadt':         stadt,
                'eingangsdatum': eingang_dt.strftime('%Y-%m-%d') if eingang_dt else None,
                'strasse':       str(strasse).strip(),
                'hausnummer':    str(cell(row, cols.get('hausnummer')) or '').strip(),
                'hausnummer_z':  str(cell(row, cols.get('hausnummer_z')) or '').strip(),
                'name':          str(cell(row, cols['name']) or '').strip(),
                'telefon':       str(cell(row, cols.get('telefon')) or '').strip(),
                'kommentar':     str(cell(row, cols.get('kommentar')) or '').strip(),
                'wartegrund':    wg_text,
                'status':        status_val,
                'genehmigung':      genehmigung,
                'genehm_von':       genehm_von.strftime('%d.%m.%Y') if genehm_von else None,
                'genehm_bis':       genehm_bis.strftime('%d.%m.%Y') if genehm_bis else None,
                'ansprechpartner':  str(cell(row, cols.get('ansprechpartner')) or '').strip(),
                'tb_termin':     tb_dt.strftime('%Y-%m-%d %H:%M') if tb_dt else None,
                'tb_termin_bis': tb_bis_dt.strftime('%Y-%m-%d') if tb_bis_dt else None,
                'tb_team':       str(cell(row, cols['tb_team']) or '').strip(),
                'tb_kommentar':  str(cell(row, cols.get('tb_kommentar')) or '').strip(),
                'inst_termin':   inst_dt.strftime('%Y-%m-%d %H:%M') if inst_dt else None,
                'inst_team':     str(cell(row, cols['inst_team']) or '').strip(),
                'inst_kommentar':str(cell(row, cols.get('inst_kommentar')) or '').strip(),
                'adresse':      full_address({
                    'strasse': str(strasse).strip(),
                    'hausnummer': cell(row, cols.get('hausnummer')),
                    'hausnummer_z': cell(row, cols.get('hausnummer_z')),
                    'stadt': stadt,
                }),
            })

    wb.close()
    return result
